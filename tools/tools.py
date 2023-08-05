import os
import time
import openpyxl
import requests


def list_all_files(root_dir):
    """
    列出文件夹下所有的文件
    :param root_dir:
    :return:
    """
    under_dir_files = []
    file_list = os.listdir(root_dir)
    for i in range(0, len(file_list)):
        path = os.path.join(root_dir, file_list[i])
        if os.path.isdir(path):
            under_dir_files.extend(list_all_files(path))
        if os.path.isfile(path):
            under_dir_files.append(path)
    return under_dir_files


def batch_analysis_dex(dex_dir):
    """
    返回目录下所有dex文件的完整路径
    :param dex_dir:
    :return:
    """
    dex_paths = []
    for root, folders, files in os.walk(dex_dir):
        for file in files:
            ext = os.path.splitext(file)[-1]
            if ext == '.dex':
                dex_file = os.path.join(root, file)
                dex_paths.append(dex_file)
    return dex_paths


def get_dependency_from_file(dependencies_file_path):
    """
    返回所有去重之后的完成gav列表
    :param dependencies_file_path:
    :return:
    """
    dependencies_list = []
    with open(dependencies_file_path, 'r') as f:
        content = f.readlines()
        for line in content:
            line = line.split('@')[0]
            line = line.replace('|', '').replace('+', '').replace('---', '').replace('\\', '').lstrip()
            if '>' in line:
                line = line.split('->')
                gav = line[0].split(':')
                final_gav = gav[0] + ':' + gav[1] + ':' + line[1].strip().replace('(*)', '').replace('(c)', '').strip()
                dependencies_list.append(final_gav)
            else:
                _gav = line.replace('(*)', '').replace('(c)', '').strip()
                dependencies_list.append(_gav)
    print('原依赖树依赖项个数：', len(dependencies_list))
    print('去重之后依赖项个数：', len(set(dependencies_list)))
    return set(dependencies_list)


def write_excel_xlsx(path, sheet_name, values):
    """
    将values数据写入path中的sheet_name
    :param path:
    :param sheet_name:
    :param values:
    :return:
    """
    index = len(values)
    workbook = openpyxl.load_workbook(path)  # 使用此方法不覆盖原有sheet表，可费劲
    sheet = workbook.create_sheet(sheet_name, 0)
    for i in range(2, index + 2):
        for j in range(1, len(values[i - 2]) + 1):
            sheet.cell(row=i, column=j, value=str(values[i - 2][j - 1]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def download_tpl(dependencies_list, output_path):
    """
    from dependencies_list download aar file or jar file to output_path
    :param dependencies_list:
    :param output_path:
    :return:
    """
    for v in set(dependencies_list):
        gav = v.split('@')
        groupId = gav[0]
        artifactId = gav[1]
        version = gav[2]
        # print(groupId, artifactId, version)
        base_url1 = 'https://maven.google.com'
        url1 = "{}/{}/{}/{}/{}-{}".format(base_url1, groupId.replace('.', '/'), artifactId, version, artifactId,
                                          version)
        dot_aar = '.aar'
        dot_jar = '.jar'
        base_url2 = 'https://repo1.maven.org/maven2'
        url2 = "{}/{}/{}/{}/{}-{}".format(base_url2, groupId.replace('.', '/'), artifactId, version, artifactId,
                                          version)
        urls = [url1 + dot_aar, url2 + dot_jar, url2 + dot_aar, url1 + dot_jar]
        for u in urls:
            try:
                res = requests.get(u)
                time.sleep(0.1)
                if res.status_code == 200:
                    if not os.path.exists(output_path):
                        os.makedirs(output_path)
                    ext = u.split('/')[-1].split('.')[-1]
                    file_name = v + '.' + ext
                    print('{}  {}'.format(file_name, u))
                    file_path = os.path.join(output_path, file_name)
                    with open(file_path, 'wb') as f:
                        f.write(res.content)
                break  # 正确url之后应该退出循环
            except Exception:
                print('异常', u)
                time.sleep(3)


def smali_to_java(smali_type):
    """
    Android smali类型转换为java类型
    :param smali_type:
    :return:
    """
    java_type = ""
    if smali_type == "Z":
        java_type = "boolean"
    elif smali_type == "B":
        java_type = "byte"
    elif smali_type == "S":
        java_type = "short"
    elif smali_type == "C":
        java_type = "char"
    elif smali_type == "I":
        java_type = "int"
    elif smali_type == "J":
        java_type = "long"
    elif smali_type == "F":
        java_type = "float"
    elif smali_type == "D":
        java_type = "double"
    elif smali_type == "V":
        java_type = "void"
    elif smali_type.startswith("["):
        # 处理数组类型
        element_type = smali_to_java(smali_type[1:])
        java_type = f"{element_type}[]"
    elif smali_type.startswith("L") and smali_type.endswith(";"):
        # 处理引用类型
        class_name = smali_type[1:-1].replace("/", ".")
        java_type = class_name

    return java_type


def format_method_keep_rule(method_name_set):
    """
    生成适用于Android R8的方法保留规则
    :param method_name_set:
    :return:
    """
    method_keep_rules = set()
    for method in method_name_set:
        class_name = str(method.class_name[1:-1]).replace('/', '.')
        method_name = method.name
        param_info = str(method.descriptor)

        index = param_info.index(')')  # 找到第一个 ')' 的索引
        param = param_info[1:index]
        params = param.split(' ')
        param_res = param
        if len(params):
            j_types = []
            for p_type in params:
                j_type = smali_to_java(p_type)
                j_types.append(j_type)
            temp = ', '.join(j_types)
            param_res = '({});'.format(temp)
        keep_rule = '-keep class %s { %s%s }' % (class_name, method_name, param_res)
        method_keep_rules.add(keep_rule)
    return method_keep_rules


def keep_rule_file_add_base_rule(base_rule_path, dependency_keep_rule_path):
    """
    将Android R8需要的默认配置添加到Android R8的保留规则文件中
    :param base_rule_path:
    :param dependency_keep_rule_path:
    :return:
    """
    with open(base_rule_path, 'r', encoding='utf-8') as f:
        data = f.readlines()
    for root, directories, files in os.walk(dependency_keep_rule_path):
        # 将文件添加到文件列表中
        for filename in files:
            file_path = os.path.join(root, filename)
            with open(file_path, 'a') as ff:
                ff.writelines(data)


def android_r8_shrink(output_path, pg_conf_path, input_path, tpl_name):
    """
    使用Android R8工具对TPL代码收缩
    :param output_path:
    :param pg_conf_path:
    :param input_path:
    :param tpl_name:
    :return:
    """
    try:
        cmd = 'java -jar  %s ' \
              '--release ' \
              '--no-minification ' \
              '--output %s ' \
              '--pg-conf %s ' \
              '--libs E:\\android\\sdk\\platforms\\android-33\\android.jar ' \
              '--libs E:\\JDK8 %s' \
              % ('../../libs/r8-3.2.74.jar', output_path, pg_conf_path, input_path)
        os.system(cmd)
    except Exception:
        print("Android R8 Error")
    # 重命名
    old_path = os.path.join(output_path, 'classes.dex')
    new_path = os.path.join(output_path, tpl_name + '.dex')
    try:
        os.rename(old_path, new_path)
    except Exception:
        os.remove(new_path)
        os.rename(old_path, new_path)


def get_filename_from_path(file_path):
    """
    给定某一文件路径，获取文件的名称，去掉后缀
    :param file_path:
    :return:
    """
    filename = file_path.split('\\')[-1]
    format_filename = os.path.splitext(filename)[0]
    return format_filename
